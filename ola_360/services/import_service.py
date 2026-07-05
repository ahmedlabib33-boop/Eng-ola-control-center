from __future__ import annotations

import csv
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from ola_360.repositories.app_repository import AppRepository


class ImportService:
    def __init__(self, repo: AppRepository):
        self.repo = repo

    def preview_csv(self, path: str) -> list[dict[str, str]]:
        return self.repo.import_csv_preview(Path(path))

    def validate_upload_name(self, file_name: str) -> bool:
        return Path(file_name).suffix.lower() in {".csv", ".xlsx"}

    def rows_from_file(self, path: str, limit: int = 500) -> list[dict[str, str]]:
        source = Path(path)
        if source.suffix.lower() == ".csv":
            return self.repo.import_csv_preview(source, limit=limit)
        if source.suffix.lower() == ".xlsx":
            workbook = load_workbook(source, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(value or "").strip() for value in rows[0]]
            data: list[dict[str, str]] = []
            for raw in rows[1 : limit + 1]:
                data.append({headers[index]: "" if value is None else str(value) for index, value in enumerate(raw) if index < len(headers) and headers[index]})
            return data
        raise ValueError("Only CSV and XLSX files are supported")

    def import_file(self, path: str, template_type: str) -> dict[str, int | list[str]]:
        rows = self.rows_from_file(path, limit=500)
        handlers: dict[str, tuple[set[str], Callable[[dict[str, str]], int]]] = {
            "warnings": (
                {"title", "category", "sector", "project", "severity", "trend", "owner", "due_date"},
                self.repo.create_warning,
            ),
            "projects": (
                {"name", "sector"},
                self.repo.create_project_from_data,
            ),
            "meetings": (
                {"title", "meeting_date"},
                self.repo.create_meeting_from_data,
            ),
            "attendees": (
                {"meeting_id", "name"},
                self.repo.create_attendee_from_data,
            ),
            "agenda_items": (
                {"meeting_id", "title"},
                self.repo.create_agenda_item_from_data,
            ),
            "warning_evidence": (
                {"warning_id", "evidence_text"},
                self.repo.create_warning_evidence_from_data,
            ),
            "commitments": (
                {"title", "owner", "due_date"},
                self.repo.create_commitment_from_data,
            ),
            "decisions": (
                {"title"},
                self.repo.create_decision_from_data,
            ),
            "milestones": (
                {"title", "due_date"},
                self.repo.create_milestone_from_data,
            ),
            "project_updates": (
                {"project", "sector", "update_date", "summary"},
                self.repo.create_project_update,
            ),
            "comments": (
                {"body"},
                self.repo.create_comment_from_data,
            ),
            "attachments": (
                {"file_path"},
                self.repo.create_attachment_from_data,
            ),
            "personal_tasks": (
                {"title"},
                self.repo.create_private_task_from_data,
            ),
            "personal_events": (
                {"title", "event_date"},
                self.repo.create_personal_event_from_data,
            ),
            "private_notes": (
                {"title"},
                self.repo.create_private_note_from_data,
            ),
            "wellbeing_checkins": (
                {"checkin_date"},
                self.repo.create_wellbeing_checkin_from_data,
            ),
        }
        if template_type not in handlers:
            raise ValueError(f"Unsupported template type: {template_type}")
        required, handler = handlers[template_type]
        errors: list[str] = []
        imported = 0
        for index, row in enumerate(rows, start=2):
            missing = sorted(key for key in required if not str(row.get(key, "")).strip())
            if missing:
                errors.append(f"Row {index}: missing {', '.join(missing)}")
                continue
            try:
                handler(row)
                imported += 1
            except Exception as exc:
                errors.append(f"Row {index}: {exc}")
        return {"imported": imported, "errors": errors}

    def import_csv(self, path: str, template_type: str) -> dict[str, int | list[str]]:
        return self.import_file(path, template_type)
